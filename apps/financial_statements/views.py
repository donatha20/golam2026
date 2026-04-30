"""
Views for financial statements generation and management.
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Sum
from django.core.paginator import Paginator
from datetime import datetime, timedelta

from .models import (
    AccountingPeriod, FinancialStatementRun, AccountClassification,
    BudgetPeriod, BudgetLine
)
from .forms import (
    FinancialStatementGenerationForm, AccountingPeriodForm, AccountClassificationForm,
    BudgetPeriodForm, BudgetLineForm, PeriodFilterForm
)
from .services import FinancialStatementGenerator
from apps.accounts.models import UserActivity

# Import export utilities
from apps.core.utils.export_utils import export_financial_statement_pdf, ExcelExporter


@login_required
def dashboard(request):
    """Financial statements dashboard."""
    # Get recent statement runs
    recent_runs = FinancialStatementRun.objects.select_related('period').order_by('-created_at')[:10]
    
    # Get current period
    current_period = AccountingPeriod.objects.filter(
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date(),
        status='open'
    ).first()
    
    # Statistics
    stats = {
        'total_periods': AccountingPeriod.objects.count(),
        'open_periods': AccountingPeriod.objects.filter(status='open').count(),
        'closed_periods': AccountingPeriod.objects.filter(status='closed').count(),
        'total_runs': FinancialStatementRun.objects.count(),
        'recent_runs_count': FinancialStatementRun.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=30)
        ).count(),
    }
    
    # Statement type breakdown
    statement_stats = FinancialStatementRun.objects.values('statement_type').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'recent_runs': recent_runs,
        'current_period': current_period,
        'stats': stats,
        'statement_stats': statement_stats,
        'title': 'Financial Statements Dashboard',
        'page_title': 'Financial Statements',
    }
    
    return render(request, 'financial_statements/dashboard.html', context)


@login_required
def generate_statement(request):
    """Generate financial statements."""
    if request.method == 'POST':
        form = FinancialStatementGenerationForm(request.POST)
        if form.is_valid():
            statement_type = form.cleaned_data['statement_type']
            period = form.cleaned_data['period']
            comparison_period = form.cleaned_data.get('comparison_period')
            export_format = form.cleaned_data['export_format']
            
            try:
                # Generate the statement
                if statement_type == 'trial_balance':
                    run, data = FinancialStatementGenerator.generate_trial_balance(
                        period=period, user=request.user
                    )
                elif statement_type == 'balance_sheet':
                    run, data = FinancialStatementGenerator.generate_balance_sheet(
                        period=period, comparison_period=comparison_period, user=request.user
                    )
                elif statement_type == 'income_statement':
                    run, data = FinancialStatementGenerator.generate_income_statement(
                        period=period, comparison_period=comparison_period, user=request.user
                    )
                elif statement_type == 'cash_flow':
                    run, data = FinancialStatementGenerator.generate_cash_flow_statement(
                        period=period, user=request.user
                    )
                elif statement_type == 'complete_set':
                    run, data = FinancialStatementGenerator.generate_complete_financial_statements(
                        period=period, comparison_period=comparison_period, user=request.user
                    )
                
                # Log activity
                UserActivity.objects.create(
                    user=request.user,
                    action='CREATE',
                    model_name='FinancialStatementRun',
                    object_id=run.id,
                    description=f'Generated {run.get_statement_type_display()} for {period.name}',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
                )
                
                messages.success(request, f'{run.get_statement_type_display()} generated successfully!')
                
                # Redirect based on export format
                if export_format == 'html':
                    return redirect('financial_statements:view_statement', run_id=run.id)
                elif export_format == 'pdf':
                    return redirect('financial_statements:export_pdf', run_id=run.id)
                elif export_format == 'excel':
                    return redirect('financial_statements:export_excel', run_id=run.id)
                
            except Exception as e:
                messages.error(request, f'Error generating statement: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = FinancialStatementGenerationForm()
    
    context = {
        'form': form,
        'title': 'Generate Financial Statement',
        'page_title': 'Generate Statement',
    }
    
    return render(request, 'financial_statements/generate_statement.html', context)


@login_required
def view_statement(request, run_id):
    """View a generated financial statement."""
    run = get_object_or_404(FinancialStatementRun, id=run_id)
    
    context = {
        'run': run,
        'data': run.results,
        'title': f'{run.get_statement_type_display()} - {run.period.name}',
        'page_title': run.get_statement_type_display(),
    }
    
    # Use different templates based on statement type
    template_map = {
        'trial_balance': 'financial_statements/trial_balance.html',
        'balance_sheet': 'financial_statements/balance_sheet.html',
        'income_statement': 'financial_statements/income_statement.html',
        'cash_flow': 'financial_statements/cash_flow.html',
        'complete_set': 'financial_statements/complete_statements.html',
    }
    
    template = template_map.get(run.statement_type, 'financial_statements/statement_view.html')
    return render(request, template, context)


@login_required
def statement_runs(request):
    """List all financial statement runs."""
    filter_form = PeriodFilterForm(request.GET)
    
    runs = FinancialStatementRun.objects.select_related('period', 'comparison_period', 'created_by').order_by('-created_at')
    
    # Apply filters
    if filter_form.is_valid():
        period = filter_form.cleaned_data.get('period')
        if period:
            runs = runs.filter(period=period)
        
        status = filter_form.cleaned_data.get('status')
        if status:
            runs = runs.filter(status=status)
        
        year = filter_form.cleaned_data.get('year')
        if year:
            runs = runs.filter(period__start_date__year=year)
    
    # Pagination
    paginator = Paginator(runs, 25)
    page_number = request.GET.get('page')
    runs = paginator.get_page(page_number)
    
    context = {
        'runs': runs,
        'filter_form': filter_form,
        'title': 'Financial Statement Runs',
        'page_title': 'Statement History',
    }
    
    return render(request, 'financial_statements/statement_runs.html', context)


@login_required
def manage_periods(request):
    """Manage accounting periods."""
    periods = AccountingPeriod.objects.order_by('-start_date')
    
    # Pagination
    paginator = Paginator(periods, 25)
    page_number = request.GET.get('page')
    periods = paginator.get_page(page_number)
    
    context = {
        'periods': periods,
        'title': 'Manage Accounting Periods',
        'page_title': 'Accounting Periods',
    }
    
    return render(request, 'financial_statements/manage_periods.html', context)


@login_required
def add_period(request):
    """Add new accounting period."""
    if request.method == 'POST':
        form = AccountingPeriodForm(request.POST)
        if form.is_valid():
            period = form.save(commit=False)
            period.created_by = request.user
            period.save()
            
            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='CREATE',
                model_name='AccountingPeriod',
                object_id=period.id,
                description=f'Created accounting period: {period.name}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
            )
            
            messages.success(request, f'Accounting period "{period.name}" created successfully!')
            return redirect('financial_statements:manage_periods')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AccountingPeriodForm()
    
    context = {
        'form': form,
        'title': 'Add Accounting Period',
        'page_title': 'New Period',
    }
    
    return render(request, 'financial_statements/add_period.html', context)


@login_required
def close_period(request, period_id):
    """Close an accounting period."""
    period = get_object_or_404(AccountingPeriod, id=period_id)
    
    if request.method == 'POST':
        if period.status == 'open':
            period.status = 'closed'
            period.closed_by = request.user
            period.closed_at = timezone.now()
            period.save()
            
            # Log activity
            UserActivity.objects.create(
                user=request.user,
                action='UPDATE',
                model_name='AccountingPeriod',
                object_id=period.id,
                description=f'Closed accounting period: {period.name}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
            )
            
            messages.success(request, f'Period "{period.name}" closed successfully!')
        else:
            messages.error(request, 'Only open periods can be closed.')
        
        return redirect('financial_statements:manage_periods')
    
    context = {
        'period': period,
        'title': f'Close Period - {period.name}',
        'page_title': 'Close Period',
    }
    
    return render(request, 'financial_statements/close_period.html', context)


@login_required
def manage_classifications(request):
    """Manage account classifications."""
    classifications = AccountClassification.objects.select_related('account').order_by('classification_type', 'sort_order')
    
    # Pagination
    paginator = Paginator(classifications, 25)
    page_number = request.GET.get('page')
    classifications = paginator.get_page(page_number)
    
    context = {
        'classifications': classifications,
        'title': 'Manage Account Classifications',
        'page_title': 'Account Classifications',
    }
    
    return render(request, 'financial_statements/manage_classifications.html', context)


@login_required
def add_classification(request):
    """Add account classification."""
    if request.method == 'POST':
        form = AccountClassificationForm(request.POST)
        if form.is_valid():
            classification = form.save(commit=False)
            classification.created_by = request.user
            classification.save()
            
            messages.success(request, 'Account classification added successfully!')
            return redirect('financial_statements:manage_classifications')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AccountClassificationForm()
    
    context = {
        'form': form,
        'title': 'Add Account Classification',
        'page_title': 'New Classification',
    }
    
    return render(request, 'financial_statements/add_classification.html', context)


@login_required
def export_pdf(request, run_id):
    """Export financial statement as PDF."""
    run = get_object_or_404(FinancialStatementRun, id=run_id)

    try:
        # Get statement data
        statement_data = run.statement_data
        statement_type = run.get_statement_type_display()
        period = f"{run.period.start_date} to {run.period.end_date}"

        # Generate PDF
        return export_financial_statement_pdf(statement_data, statement_type, period)

    except Exception as e:
        messages.error(request, f'Error generating PDF: {str(e)}')
        return redirect('financial_statements:view_statement', run_id=run_id)


@login_required
def export_excel(request, run_id):
    """Export financial statement as Excel."""
    run = get_object_or_404(FinancialStatementRun, id=run_id)

    try:
        # Get statement data
        statement_data = run.statement_data
        statement_type = run.get_statement_type_display()
        period = f"{run.period.start_date} to {run.period.end_date}"

        # Prepare data for Excel export
        exporter = ExcelExporter(title=f"{statement_type} - {period}")
        filename = f"{statement_type.replace(' ', '_').lower()}_{run.period.start_date}_{run.period.end_date}.xlsx"

        if run.statement_type == 'trial_balance':
            headers = ['Account Code', 'Account Name', 'Debit Balance', 'Credit Balance']
            data = []
            for item in statement_data.get('accounts', []):
                data.append([
                    item.get('account_code', ''),
                    item.get('account_name', ''),
                    item.get('debit_balance', 0),
                    item.get('credit_balance', 0)
                ])

            # Add totals row
            data.append([
                '', 'TOTALS',
                statement_data.get('total_debits', 0),
                statement_data.get('total_credits', 0)
            ])

        elif run.statement_type == 'balance_sheet':
            headers = ['Item', 'Current Period', 'Previous Period', 'Change']
            data = []

            # Assets
            data.append(['ASSETS', '', '', ''])
            for asset in statement_data.get('assets', []):
                data.append([
                    f"  {asset.get('name', '')}",
                    asset.get('current_amount', 0),
                    asset.get('previous_amount', 0),
                    asset.get('change', 0)
                ])

            # Liabilities
            data.append(['LIABILITIES', '', '', ''])
            for liability in statement_data.get('liabilities', []):
                data.append([
                    f"  {liability.get('name', '')}",
                    liability.get('current_amount', 0),
                    liability.get('previous_amount', 0),
                    liability.get('change', 0)
                ])

            # Equity
            data.append(['EQUITY', '', '', ''])
            for equity in statement_data.get('equity', []):
                data.append([
                    f"  {equity.get('name', '')}",
                    equity.get('current_amount', 0),
                    equity.get('previous_amount', 0),
                    equity.get('change', 0)
                ])

        elif run.statement_type == 'income_statement':
            headers = ['Item', 'Current Period', 'Previous Period', 'Change %']
            data = []

            # Revenue
            data.append(['REVENUE', '', '', ''])
            for revenue in statement_data.get('revenue', []):
                data.append([
                    f"  {revenue.get('name', '')}",
                    revenue.get('current_amount', 0),
                    revenue.get('previous_amount', 0),
                    revenue.get('change_percentage', 0)
                ])

            # Expenses
            data.append(['EXPENSES', '', '', ''])
            for expense in statement_data.get('expenses', []):
                data.append([
                    f"  {expense.get('name', '')}",
                    expense.get('current_amount', 0),
                    expense.get('previous_amount', 0),
                    expense.get('change_percentage', 0)
                ])

        else:
            # Default format
            headers = ['Description', 'Amount']
            data = [[item.get('description', ''), item.get('amount', 0)] for item in statement_data]

        title = f"{statement_type} - {period}"
        return exporter.generate_table_excel(data, headers, filename, title, statement_type)

    except Exception as e:
        messages.error(request, f'Error generating Excel file: {str(e)}')
        return redirect('financial_statements:view_statement', run_id=run_id)


@login_required
def financial_summary_api(request):
    """API endpoint for financial summary data."""
    # Get current period
    current_period = AccountingPeriod.objects.filter(
        start_date__lte=timezone.now().date(),
        end_date__gte=timezone.now().date(),
        status='open'
    ).first()
    
    if not current_period:
        return JsonResponse({'error': 'No current period found'}, status=404)
    
    try:
        # Generate quick summary data
        service = FinancialStatementGenerator()
        _, trial_balance = service.generate_trial_balance(current_period)
        _, balance_sheet = service.generate_balance_sheet(current_period)
        
        summary_data = {
            'period': {
                'name': current_period.name,
                'start_date': current_period.start_date.isoformat(),
                'end_date': current_period.end_date.isoformat(),
            },
            'trial_balance': {
                'total_debits': float(trial_balance['total_debits']),
                'total_credits': float(trial_balance['total_credits']),
                'is_balanced': trial_balance['is_balanced'],
            },
            'balance_sheet': {
                'total_assets': float(balance_sheet['assets']['total_assets']),
                'total_liabilities': float(balance_sheet['liabilities']['total_liabilities']),
                'total_equity': float(balance_sheet['equity']['total_equity']),
                'is_balanced': balance_sheet['is_balanced'],
            }
        }
        
        return JsonResponse(summary_data)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ============ SPECIFIC FINANCIAL REPORTS ============

@login_required
def trial_balance_report(request):
    """Generate trial balance report."""
    from apps.accounting.models import Account, JournalEntry, JournalEntryLine
    from decimal import Decimal

    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Get all accounts with their balances
    accounts = Account.objects.all().order_by('account_code')
    trial_balance_data = []
    total_debits = Decimal('0.00')
    total_credits = Decimal('0.00')

    for account in accounts:
        # Calculate account balance for the period using journal entry lines
        journal_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__entry_date__range=[start_date, end_date],
            journal_entry__is_posted=True
        )

        debit_total = journal_lines.aggregate(
            total=Sum('debit_amount')
        )['total'] or Decimal('0.00')

        credit_total = journal_lines.aggregate(
            total=Sum('credit_amount')
        )['total'] or Decimal('0.00')

        balance = debit_total - credit_total

        if balance != 0:  # Only include accounts with balances
            if balance > 0:
                debit_balance = balance
                credit_balance = Decimal('0.00')
                total_debits += debit_balance
            else:
                debit_balance = Decimal('0.00')
                credit_balance = abs(balance)
                total_credits += credit_balance

            trial_balance_data.append({
                'account': account,
                'debit_balance': debit_balance,
                'credit_balance': credit_balance,
            })

    context = {
        'trial_balance_data': trial_balance_data,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'difference': total_debits - total_credits,
        'start_date': start_date,
        'end_date': end_date,
        'is_balanced': total_debits == total_credits,
    }

    return render(request, 'financial_statements/trial_balance_report.html', context)


@login_required
def balance_sheet_report(request):
    """Generate balance sheet report."""
    from apps.accounting.models import Account, JournalEntryLine
    from decimal import Decimal

    # Get date from request
    as_of_date = request.GET.get('as_of_date')
    if not as_of_date:
        as_of_date = timezone.now().date()
    else:
        as_of_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()

    # Get accounts by category
    assets = Account.objects.filter(account_type='asset').order_by('account_code')
    liabilities = Account.objects.filter(account_type='liability').order_by('account_code')
    equity = Account.objects.filter(account_type='equity').order_by('account_code')

    # Calculate balances (simplified - you may need to adjust based on your accounting logic)
    total_assets = Decimal('0.00')
    total_liabilities = Decimal('0.00')
    total_equity = Decimal('0.00')

    context = {
        'assets': assets,
        'liabilities': liabilities,
        'equity': equity,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'as_of_date': as_of_date,
    }

    return render(request, 'financial_statements/balance_sheet_report.html', context)


@login_required
def income_statement_report(request):
    """Generate income statement report."""
    from apps.finance_tracker.models import Income, Expenditure
    from decimal import Decimal

    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Get income data
    income_data = Income.objects.filter(
        income_date__range=[start_date, end_date],
        status='RECEIVED'
    ).values('source').annotate(
        total=Sum('amount')
    ).order_by('source')

    # Get expenditure data
    expenditure_data = Expenditure.objects.filter(
        expenditure_date__range=[start_date, end_date],
        status='paid'
    ).values('expenditure_type').annotate(
        total=Sum('amount')
    ).order_by('expenditure_type')

    # Calculate totals
    total_income = Income.objects.filter(
        income_date__range=[start_date, end_date],
        status='RECEIVED'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    total_expenses = Expenditure.objects.filter(
        expenditure_date__range=[start_date, end_date],
        status='paid'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    net_income = total_income - total_expenses

    context = {
        'income_data': income_data,
        'expenditure_data': expenditure_data,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_income': net_income,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'financial_statements/income_statement_report.html', context)


@login_required
def cash_flow_report(request):
    """Generate cash flow statement."""
    # This is a simplified version - you may need to enhance based on your requirements
    context = {
        'title': 'Cash Flow Statement',
        'message': 'Cash flow report functionality will be implemented based on your specific requirements.',
    }
    return render(request, 'financial_statements/cash_flow_report.html', context)


@login_required
def portfolio_analysis_report(request):
    """Generate portfolio analysis report."""
    from apps.loans.models import Loan
    from decimal import Decimal

    # Get loan portfolio statistics
    total_loans = Loan.objects.count()
    active_loans = Loan.objects.filter(status='active').count()

    # Portfolio by status
    portfolio_by_status = Loan.objects.values('status').annotate(
        count=Count('id'),
        total_amount=Sum('amount_approved')
    ).order_by('status')

    # Portfolio by loan category
    portfolio_by_category = Loan.objects.values('loan_category').annotate(
        count=Count('id'),
        total_amount=Sum('amount_approved')
    ).order_by('loan_category')

    # Calculate percentages for portfolio by status
    for item in portfolio_by_status:
        if total_loans > 0:
            item['percentage'] = round((item['count'] / total_loans) * 100, 1)
        else:
            item['percentage'] = 0

    # Calculate percentages for portfolio by category
    for item in portfolio_by_category:
        if total_loans > 0:
            item['percentage'] = round((item['count'] / total_loans) * 100, 1)
        else:
            item['percentage'] = 0

    # Calculate active loan ratio
    if total_loans > 0:
        active_loan_ratio = round((active_loans / total_loans) * 100, 1)
    else:
        active_loan_ratio = 0

    context = {
        'total_loans': total_loans,
        'active_loans': active_loans,
        'active_loan_ratio': active_loan_ratio,
        'portfolio_by_status': portfolio_by_status,
        'portfolio_by_category': portfolio_by_category,
    }

    return render(request, 'financial_statements/portfolio_analysis_report.html', context)


@login_required
def loan_aging_report(request):
    """Generate loan aging report."""
    from apps.loans.models import Loan
    from apps.repayments.models import LoanRepaymentSchedule
    from datetime import timedelta

    today = timezone.now().date()

    # Define aging buckets
    aging_buckets = {
        'current': {'min': 0, 'max': 0, 'loans': []},
        '1_30_days': {'min': 1, 'max': 30, 'loans': []},
        '31_60_days': {'min': 31, 'max': 60, 'loans': []},
        '61_90_days': {'min': 61, 'max': 90, 'loans': []},
        'over_90_days': {'min': 91, 'max': 999, 'loans': []},
    }

    # Get overdue loans
    overdue_schedules = LoanRepaymentSchedule.objects.filter(
        due_date__lt=today,
        is_paid=False
    ).select_related('loan')

    for schedule in overdue_schedules:
        days_overdue = (today - schedule.due_date).days

        # Categorize into aging buckets
        if days_overdue == 0:
            aging_buckets['current']['loans'].append(schedule)
        elif 1 <= days_overdue <= 30:
            aging_buckets['1_30_days']['loans'].append(schedule)
        elif 31 <= days_overdue <= 60:
            aging_buckets['31_60_days']['loans'].append(schedule)
        elif 61 <= days_overdue <= 90:
            aging_buckets['61_90_days']['loans'].append(schedule)
        else:
            aging_buckets['over_90_days']['loans'].append(schedule)

    context = {
        'aging_buckets': aging_buckets,
        'report_date': today,
    }

    return render(request, 'financial_statements/loan_aging_report.html', context)


@login_required
def collection_summary_report(request):
    """Generate collection summary report."""
    from apps.repayments.models import Payment
    from decimal import Decimal

    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date or not end_date:
        # Default to current month
        today = timezone.now().date()
        start_date = today.replace(day=1)
        end_date = today
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

    # Get collection data
    collections = Payment.objects.filter(
        payment_date__range=[start_date, end_date],
        status='completed'
    )

    # Summary statistics
    total_collections = collections.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    collection_count = collections.count()

    # Collections by payment method
    collections_by_method = collections.values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('payment_method')

    # Daily collections with calculated averages
    daily_collections = collections.values('payment_date').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('payment_date')
    
    # Add average calculation to each day
    daily_collections_with_avg = []
    for day in daily_collections:
        if day['count'] > 0:
            day['average'] = day['total'] / day['count']
        else:
            day['average'] = Decimal('0.00')
        daily_collections_with_avg.append(day)
    
    # Calculate overall average
    if collection_count > 0:
        overall_average = total_collections / collection_count
    else:
        overall_average = Decimal('0.00')

    # Handle Excel export
    if request.GET.get('export') == 'excel':
        try:
            return export_collection_summary_excel(
                daily_collections_with_avg,
                collections_by_method,
                {
                    'total_collections': total_collections,
                    'collection_count': collection_count,
                    'overall_average': overall_average,
                    'start_date': start_date,
                    'end_date': end_date,
                }
            )
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            return export_collection_summary_csv(
                daily_collections_with_avg,
                collections_by_method,
                {
                    'total_collections': total_collections,
                    'collection_count': collection_count,
                    'overall_average': overall_average,
                    'start_date': start_date,
                    'end_date': end_date,
                }
            )

    context = {
        'total_collections': total_collections,
        'collection_count': collection_count,
        'collections_by_method': collections_by_method,
        'daily_collections': daily_collections_with_avg,
        'overall_average': overall_average,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'financial_statements/collection_summary_report.html', context)


def export_collection_summary_excel(daily_collections, collections_by_method, summary_data):
    """Export collection summary report to Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from io import BytesIO
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection Summary"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B7A76", end_color="2B7A76", fill_type="solid")
    title_font = Font(bold=True, size=16)
    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Collection Summary Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Date range
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Period: {summary_data['start_date']} to {summary_data['end_date']}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Summary statistics
    current_row = 4
    ws[f'A{current_row}'] = "Summary Statistics"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    ws[f'A{current_row}'] = "Total Collections:"
    ws[f'B{current_row}'] = float(summary_data['total_collections'])
    ws[f'B{current_row}'].number_format = '#,##0.00'
    current_row += 1
    
    ws[f'A{current_row}'] = "Number of Payments:"
    ws[f'B{current_row}'] = summary_data['collection_count']
    current_row += 1
    
    ws[f'A{current_row}'] = "Average Payment:"
    ws[f'B{current_row}'] = float(summary_data['overall_average'])
    ws[f'B{current_row}'].number_format = '#,##0.00'
    current_row += 2
    
    # Collections by Payment Method
    ws[f'A{current_row}'] = "Collections by Payment Method"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    # Headers for payment method table
    headers = ['Payment Method', 'Count', 'Total Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    current_row += 1
    
    # Payment method data
    for method in collections_by_method:
        ws[f'A{current_row}'] = method['payment_method'] or 'Cash'
        ws[f'B{current_row}'] = method['count']
        ws[f'C{current_row}'] = float(method['total'])
        ws[f'C{current_row}'].number_format = '#,##0.00'
        
        # Apply borders
        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1
    
    current_row += 1
    
    # Daily Collections
    ws[f'A{current_row}'] = "Daily Collections Detail"
    ws[f'A{current_row}'].font = bold_font
    current_row += 1
    
    # Headers for daily collections table
    headers = ['Date', 'Number of Payments', 'Total Amount', 'Average Payment']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    current_row += 1
    
    # Daily collections data
    for day in daily_collections:
        ws[f'A{current_row}'] = day['payment_date']
        ws[f'B{current_row}'] = day['count']
        ws[f'C{current_row}'] = float(day['total'])
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'D{current_row}'] = float(day['average'])
        ws[f'D{current_row}'].number_format = '#,##0.00'
        
        # Apply borders
        for col in range(1, 5):
            ws.cell(row=current_row, column=col).border = border
        current_row += 1
    
    # Total row
    ws[f'A{current_row}'] = "TOTAL"
    ws[f'A{current_row}'].font = bold_font
    ws[f'B{current_row}'] = summary_data['collection_count']
    ws[f'B{current_row}'].font = bold_font
    ws[f'C{current_row}'] = float(summary_data['total_collections'])
    ws[f'C{current_row}'].number_format = '#,##0.00'
    ws[f'C{current_row}'].font = bold_font
    ws[f'D{current_row}'] = float(summary_data['overall_average'])
    ws[f'D{current_row}'].number_format = '#,##0.00'
    ws[f'D{current_row}'].font = bold_font
    
    # Apply borders to total row
    for col in range(1, 5):
        ws.cell(row=current_row, column=col).border = border
    
    # Set fixed column widths to avoid merged cell issues
    ws.column_dimensions['A'].width = 20  # Date/Description column
    ws.column_dimensions['B'].width = 18  # Count column
    ws.column_dimensions['C'].width = 15  # Amount column
    ws.column_dimensions['D'].width = 15  # Average column
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"collection_summary_{summary_data['start_date']}_{summary_data['end_date']}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def export_collection_summary_csv(daily_collections, collections_by_method, summary_data):
    """Export collection summary report to CSV as fallback."""
    import csv
    from django.http import HttpResponse
    from io import StringIO
    
    output = StringIO()
    writer = csv.writer(output)
    
    # Title and header
    writer.writerow(['Collection Summary Report'])
    writer.writerow([f"Period: {summary_data['start_date']} to {summary_data['end_date']}"])
    writer.writerow([])  # Empty row
    
    # Summary statistics
    writer.writerow(['Summary Statistics'])
    writer.writerow(['Total Collections:', f"{summary_data['total_collections']:.2f}"])
    writer.writerow(['Number of Payments:', summary_data['collection_count']])
    writer.writerow(['Average Payment:', f"{summary_data['overall_average']:.2f}"])
    writer.writerow([])  # Empty row
    
    # Collections by Payment Method
    writer.writerow(['Collections by Payment Method'])
    writer.writerow(['Payment Method', 'Count', 'Total Amount'])
    for method in collections_by_method:
        writer.writerow([
            method['payment_method'] or 'Cash',
            method['count'],
            f"{method['total']:.2f}"
        ])
    writer.writerow([])  # Empty row
    
    # Daily Collections
    writer.writerow(['Daily Collections Detail'])
    writer.writerow(['Date', 'Number of Payments', 'Total Amount', 'Average Payment'])
    for day in daily_collections:
        writer.writerow([
            day['payment_date'],
            day['count'],
            f"{day['total']:.2f}",
            f"{day['average']:.2f}"
        ])
    
    # Total row
    writer.writerow([
        'TOTAL',
        summary_data['collection_count'],
        f"{summary_data['total_collections']:.2f}",
        f"{summary_data['overall_average']:.2f}"
    ])
    
    # Create response
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    filename = f"collection_summary_{summary_data['start_date']}_{summary_data['end_date']}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


