"""
Export utilities for PDF and Excel generation.
"""
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.utils.dataframe import dataframe_to_rows  # Not needed
# import pandas as pd  # Not needed


class PDFExporter:
    """Utility class for PDF export functionality."""
    
    def __init__(self, title="Report", author="Golam Financial Services"):
        self.title = title
        self.author = author
        self.styles = getSampleStyleSheet()
        self.custom_styles = self._create_custom_styles()
    
    def _create_custom_styles(self):
        """Create custom paragraph styles."""
        styles = {}
        
        # Title style
        styles['CustomTitle'] = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.HexColor('#2b7a76'),
            alignment=1  # Center alignment
        )
        
        # Heading style
        styles['CustomHeading'] = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#2b7a76')
        )
        
        # Normal style with custom color
        styles['CustomNormal'] = ParagraphStyle(
            'CustomNormal',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.black
        )
        
        return styles
    
    def create_pdf_response(self, filename):
        """Create HTTP response for PDF download."""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def generate_table_pdf(self, data, headers, filename, title=None):
        """Generate PDF with table data."""
        response = self.create_pdf_response(filename)
        buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Build content
        story = []
        
        # Add title
        if title:
            title_para = Paragraph(title, self.custom_styles['CustomTitle'])
            story.append(title_para)
            story.append(Spacer(1, 12))
        
        # Add generation info
        generation_info = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        info_para = Paragraph(generation_info, self.custom_styles['CustomNormal'])
        story.append(info_para)
        story.append(Spacer(1, 20))
        
        # Prepare table data
        table_data = [headers]
        for row in data:
            formatted_row = []
            for cell in row:
                if isinstance(cell, Decimal):
                    formatted_row.append(f"₹{cell:,.2f}")
                elif isinstance(cell, datetime):
                    formatted_row.append(cell.strftime('%Y-%m-%d'))
                else:
                    formatted_row.append(str(cell) if cell is not None else '')
            table_data.append(formatted_row)
        
        # Create table
        table = Table(table_data)
        
        # Style the table
        table_style = TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b7a76')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ])
        
        table.setStyle(table_style)
        story.append(table)
        
        # Build PDF
        doc.build(story)
        
        # Get PDF data
        pdf_data = buffer.getvalue()
        buffer.close()
        
        response.write(pdf_data)
        return response
    
    def generate_financial_statement_pdf(self, statement_data, statement_type, period):
        """Generate PDF for financial statements."""
        filename = f"{statement_type.replace(' ', '_').lower()}_{period}.pdf"
        title = f"{statement_type} - {period}"
        
        # Prepare data based on statement type
        if statement_type == "Trial Balance":
            headers = ['Account', 'Debit', 'Credit']
            data = []
            for item in statement_data:
                data.append([
                    item.get('account_name', ''),
                    item.get('debit_balance', Decimal('0.00')),
                    item.get('credit_balance', Decimal('0.00'))
                ])
        
        elif statement_type == "Balance Sheet":
            headers = ['Item', 'Amount']
            data = []
            # Assets
            data.append(['ASSETS', ''])
            for asset in statement_data.get('assets', []):
                data.append([f"  {asset.get('name', '')}", asset.get('amount', Decimal('0.00'))])
            
            # Liabilities
            data.append(['LIABILITIES', ''])
            for liability in statement_data.get('liabilities', []):
                data.append([f"  {liability.get('name', '')}", liability.get('amount', Decimal('0.00'))])
            
            # Equity
            data.append(['EQUITY', ''])
            for equity in statement_data.get('equity', []):
                data.append([f"  {equity.get('name', '')}", equity.get('amount', Decimal('0.00'))])
        
        else:
            # Default format
            headers = ['Description', 'Amount']
            data = [[item.get('description', ''), item.get('amount', Decimal('0.00'))] for item in statement_data]
        
        return self.generate_table_pdf(data, headers, filename, title)


class ExcelExporter:
    """Utility class for Excel export functionality."""
    
    def __init__(self, title="Report"):
        self.title = title
        self.workbook = None
        self.worksheet = None
    
    def create_excel_response(self, filename):
        """Create HTTP response for Excel download."""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _apply_header_style(self, worksheet, row_num, col_count):
        """Apply styling to header row."""
        # Header font and fill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B7A76", end_color="2B7A76", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply to header row
        for col in range(1, col_count + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _apply_data_style(self, worksheet, start_row, end_row, col_count):
        """Apply styling to data rows."""
        # Data font and alignment
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply to data rows
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    def generate_table_excel(self, data, headers, filename, title=None, sheet_name="Data"):
        """Generate Excel file with table data."""
        response = self.create_excel_response(filename)
        
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        current_row = 1
        
        # Add title if provided
        if title:
            worksheet.cell(row=current_row, column=1, value=title)
            title_cell = worksheet.cell(row=current_row, column=1)
            title_cell.font = Font(size=16, bold=True, color="2B7A76")
            current_row += 2
        
        # Add generation info
        generation_info = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        worksheet.cell(row=current_row, column=1, value=generation_info)
        current_row += 2
        
        # Add headers
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=current_row, column=col, value=header)
        
        self._apply_header_style(worksheet, current_row, len(headers))
        header_row = current_row
        current_row += 1
        
        # Add data
        for row_data in data:
            for col, value in enumerate(row_data, 1):
                if isinstance(value, Decimal):
                    worksheet.cell(row=current_row, column=col, value=float(value))
                elif isinstance(value, datetime):
                    worksheet.cell(row=current_row, column=col, value=value)
                else:
                    worksheet.cell(row=current_row, column=col, value=value)
            current_row += 1
        
        # Apply data styling
        if current_row > header_row + 1:
            self._apply_data_style(worksheet, header_row + 1, current_row - 1, len(headers))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        workbook.save(response)
        return response
    
    def generate_multi_sheet_excel(self, sheets_data, filename):
        """Generate Excel file with multiple sheets."""
        response = self.create_excel_response(filename)
        
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        for sheet_info in sheets_data:
            sheet_name = sheet_info['name']
            headers = sheet_info['headers']
            data = sheet_info['data']
            title = sheet_info.get('title', sheet_name)
            
            # Create worksheet
            worksheet = workbook.create_sheet(title=sheet_name)
            
            current_row = 1
            
            # Add title
            worksheet.cell(row=current_row, column=1, value=title)
            title_cell = worksheet.cell(row=current_row, column=1)
            title_cell.font = Font(size=14, bold=True, color="2B7A76")
            current_row += 2
            
            # Add headers
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=current_row, column=col, value=header)
            
            self._apply_header_style(worksheet, current_row, len(headers))
            header_row = current_row
            current_row += 1
            
            # Add data
            for row_data in data:
                for col, value in enumerate(row_data, 1):
                    if isinstance(value, Decimal):
                        worksheet.cell(row=current_row, column=col, value=float(value))
                    elif isinstance(value, datetime):
                        worksheet.cell(row=current_row, column=col, value=value)
                    else:
                        worksheet.cell(row=current_row, column=col, value=value)
                current_row += 1
            
            # Apply styling
            if current_row > header_row + 1:
                self._apply_data_style(worksheet, header_row + 1, current_row - 1, len(headers))
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        workbook.save(response)
        return response


# Convenience functions
def export_to_pdf(data, headers, filename, title=None):
    """Quick function to export data to PDF."""
    exporter = PDFExporter()
    return exporter.generate_table_pdf(data, headers, filename, title)


def export_to_excel(data, headers, filename, title=None, sheet_name="Data"):
    """Quick function to export data to Excel."""
    exporter = ExcelExporter()
    return exporter.generate_table_excel(data, headers, filename, title, sheet_name)


def export_financial_statement_pdf(statement_data, statement_type, period):
    """Quick function to export financial statements to PDF."""
    exporter = PDFExporter()
    return exporter.generate_financial_statement_pdf(statement_data, statement_type, period)
